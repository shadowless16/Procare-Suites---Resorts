# budget_planner.py

import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# === STEP 1: Generate Budget Data ===
print("⏳ Creating personal budget spreadsheet...")

categories = ['Rent', 'Utilities', 'Groceries', 'Transportation', 'Entertainment', 'Savings']
base_date = datetime(2025, 5, 1)

data = []

for i in range(500):
    date = base_date + timedelta(days=random.randint(0, 29))
    category = random.choice(categories)
    amount = round(random.uniform(2000, 150000), 2)
    data.append([date.strftime('%Y-%m-%d'), category, amount])

df = pd.DataFrame(data, columns=['Date', 'Category', 'Amount'])

# === STEP 2: Save raw data ===
raw_file = "monthly_budget.xlsx"
df.to_excel(raw_file, index=False)

# === STEP 3: Load into openpyxl for formatting + formulas ===
wb = load_workbook(raw_file)
ws = wb.active

# Format headers
header_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
for col in range(1, 4):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.fill = header_fill

# === STEP 4: Summary Section ===
summary_row = ws.max_row + 3
ws[f"A{summary_row}"] = "Summary"
ws[f"A{summary_row}"].font = Font(bold=True)

# Total Income (fake static income values)
ws[f"A{summary_row + 1}"] = "Salary"
ws[f"B{summary_row + 1}"] = 300000
ws[f"A{summary_row + 2}"] = "Freelance"
ws[f"B{summary_row + 2}"] = 100000

ws[f"A{summary_row + 3}"] = "Total Income"
ws[f"B{summary_row + 3}"] = "=B{}+B{}".format(summary_row + 1, summary_row + 2)

# Total Expenses using SUMIF
ws[f"A{summary_row + 4}"] = "Total Expenses"
ws[f"B{summary_row + 4}"] = f'=SUM(C2:C{ws.max_row})'

# IF Logic
ws[f"A{summary_row + 5}"] = "Budget Status"
ws[f"B{summary_row + 5}"] = f'=IF(B{summary_row + 4}>B{summary_row + 3},"Warning: Expenses exceed income!","Budget is balanced")'

# CONCATENATE Summary
ws[f"A{summary_row + 6}"] = "Summary Message"
ws[f"B{summary_row + 6}"] = f'=CONCATENATE("Total Income: ", B{summary_row + 3}, ", Total Expenses: ", B{summary_row + 4})'

# TODAY Function
ws[f"A{summary_row + 7}"] = "Report Date"
ws[f"B{summary_row + 7}"] = "=TODAY()"

# IFERROR Example (Expense-to-Income Ratio)
ws[f"A{summary_row + 8}"] = "Expense/Income Ratio"
ws[f"B{summary_row + 8}"] = f'=IFERROR(B{summary_row + 4}/B{summary_row + 3}, "Error in calculation")'

# === STEP 5: Save final file ===
final_file = "final_budget_report.xlsx"
wb.save(final_file)
print(f"✅ Final budget spreadsheet saved as {final_file}")